import openpyxl
from openpyxl import Workbook


def get_row_count(path,sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook(sheet_name)
    return sheet.max_row

def get_column_count(path,sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook(sheet_name)
    return sheet.max_column

def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    cell_value = sheet.cell(row=row_number, column=column_number).value
    workbook.close()
    return cell_value

def set_cell_data(path,sheet_name,row_number,column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value.data
    workbook.save(path)

def get_data_from_excel(path,sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook("D:/Study "
                                      "Videos/SQA/Python/pythonSQAProject/SampleProject4/ExcelFiles/DemoLoginExcel.xlsx")
    sheet = workbook['LoginTest']
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(1, total_cols + 1):
            # col_list = []
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)

    return final_list

def write_data_to_excel(path, sheet_name, data, headers=None):
    workbook = openpyxl.load_workbook(path)

    # Check if sheet exists
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        # Create new sheet if it doesn't exist
        sheet = workbook.create_sheet(title=sheet_name)

        # Write header if provided
        if headers is not None:
            sheet.append(headers)

    # Write data
    for item in data:
        sheet.append(item)

    workbook.save(path)