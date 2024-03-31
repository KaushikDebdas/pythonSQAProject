'''
- first we need to create a excel file in our pc location
- we need to install openpyxl
    - pip install openpyxl
'''
import openpyxl

# for load our excel files
workbook = openpyxl.load_workbook("D:/Study Videos/SQA/Python/pythonSQAProject/PythonSeleniumPractice/UtilitiesDemo/ExcelFiles/DemoExcel.xlsx")
# from excel read excelSheetName
sheet = workbook['LoginTest']
# total rows of particular sheet we need to use max_row
total_rows = sheet.max_row
print(total_rows)
# total column of particualar sheet we need to use max_cloumn
total_col = sheet.max_column
print(total_col)

# print particular data from excel
# for this we need to do use cell
# cell(rows,cols).value needs
particular_valuefromExcel = sheet.cell(3,1).value
print("Value at Row 3, Column 1:", particular_valuefromExcel)

# entire table form excel
for r in range(1,total_rows+1):
    for c in range(1,total_col+1):
        full_table = sheet.cell(r,c).value
        print(full_table , end="         ")
    print()

