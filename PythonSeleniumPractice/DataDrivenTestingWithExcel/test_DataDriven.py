import openpyxl
import pytest

def get_data():
    # Method 1
    final_list = []
    workbook = openpyxl.load_workbook("ExcelFiles/DemoExcel.xlsx")
    sheet = workbook['LoginTest']
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2,total_rows+1):
        row_list = []
        for c in range(1,total_cols+1):
            #col_list = []
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)

    return final_list
    # Method 2
    # return [
    #     ("kaushik", "pass1"),
    #     ("saimun", "pass2"),
    #     ("shafin", "pass3")
    # ]

@pytest.mark.parametrize("username,password",get_data())
def test_login(username,password):
    print("Logged in using username: "+username+ " and password is: "+password)