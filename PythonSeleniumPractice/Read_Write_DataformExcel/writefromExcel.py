'''
- first we need to create a excel file in our pc location
- we need to install openpyxl
    - pip install openpyxl
'''
import openpyxl

# for load our excel files
workbook = openpyxl.load_workbook("ExcelFiles/DemoExcel.xlsx")
# from excel read excelSheetName
sheet = workbook['LoginTest']
# column add in excel sheet
sheet.cell(1,4).value="role"
workbook.save("ExcelFiles/DemoExcel.xlsx") # mandatory use for save the data in excel

# write data in excel
for r in range(5,11):
    for c in range(1,4):
        sheet.cell(r,c).value="kaushik"
workbook.save("ExcelFiles/DemoExcel.xlsx") # mandatory use for save the data in excel
